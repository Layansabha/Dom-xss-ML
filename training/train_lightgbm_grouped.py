from __future__ import annotations

import argparse
import ast
import gzip
import hashlib
import json
import lzma
from collections import Counter, defaultdict
from collections.abc import Iterable, Iterator, Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any, TextIO

import joblib
import numpy as np
from lightgbm import LGBMClassifier, early_stopping, log_evaluation
from openpyxl import load_workbook
from scipy import sparse
from sklearn.metrics import (
    average_precision_score,
    confusion_matrix,
    f1_score,
    precision_score,
    recall_score,
    roc_auc_score,
)

from preprocessing.security_features import (
    FEATURE_CONTRACT,
    augment_security_features,
)


RANDOM_STATE = 42
SPLIT_NAMES = ("train", "validation", "test")
EXCEL_CELL_LIMIT = 32_767
DEFAULT_CANDIDATES = (
    {"num_leaves": 31, "max_depth": 7, "min_child_samples": 50, "learning_rate": 0.05},
    {"num_leaves": 31, "max_depth": 8, "min_child_samples": 100, "learning_rate": 0.03},
    {"num_leaves": 64, "max_depth": 12, "min_child_samples": 50, "learning_rate": 0.05},
    {"num_leaves": 64, "max_depth": 15, "min_child_samples": 20, "learning_rate": 0.10},
    {"num_leaves": 127, "max_depth": 15, "min_child_samples": 50, "learning_rate": 0.03},
)


@dataclass(frozen=True)
class ParsedRecord:
    label: int
    weight: float
    dbg: str
    script_id: str
    features: dict[str, float]
    feature_hash: str
    source_file: str


@dataclass(frozen=True)
class RejectedRecord:
    reason: str
    source_file: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Train a leakage-resistant LightGBM DOM-XSS ranking model. "
            "Inputs may be CMU JSONL, JSONL.xz, JSONL.gz, or legacy XLSX files."
        )
    )
    parser.add_argument("inputs", nargs="+", type=Path)
    parser.add_argument("--output-root", type=Path, default=Path("."))
    parser.add_argument("--vocab-size", type=int, default=500)
    parser.add_argument("--min-token-count", type=int, default=5)
    parser.add_argument(
        "--target-recall",
        type=float,
        default=0.95,
        help="Validation recall target used to select the deployment threshold.",
    )
    parser.add_argument(
        "--max-train-duplicates",
        type=int,
        default=20,
        help=(
            "Maximum rows retained per identical feature bag in training. "
            "Validation and test remain unique and unseen."
        ),
    )
    parser.add_argument("--max-rows", type=int)
    parser.add_argument(
        "--dataset-url",
        default=(
            "https://kilthub.cmu.edu/articles/dataset/"
            "DOM_XSS_Web_Vulnerability_Dataset/13870256"
        ),
    )
    parser.add_argument(
        "--paper-url",
        default=(
            "https://www.contrib.andrew.cmu.edu/~liminjia/research/"
            "papers/www2021-dom-xss-dnn.pdf"
        ),
    )
    return parser.parse_args()


def normalize_token(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    token = value.lower().strip()
    if len(token) < 3 or token.isnumeric():
        return None
    if not any(character.isalnum() for character in token):
        return None
    return token


def normalize_feature_mapping(value: object) -> dict[str, float] | None:
    if isinstance(value, str):
        try:
            value = ast.literal_eval(value)
        except (SyntaxError, ValueError):
            return None
    if not isinstance(value, Mapping):
        return None

    normalized: dict[str, float] = {}
    for raw_token, raw_count in value.items():
        token = normalize_token(raw_token)
        if token is None:
            continue
        try:
            count = float(raw_count)
        except (TypeError, ValueError):
            return None
        if not np.isfinite(count) or count < 0:
            return None
        if count:
            normalized[token] = normalized.get(token, 0.0) + count
    return augment_security_features(normalized)


def canonical_feature_hash(features: Mapping[str, float]) -> str:
    payload = json.dumps(
        sorted(features.items()),
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def split_for_script(script_id: str) -> int:
    digest = hashlib.sha256(script_id.encode("utf-8")).digest()
    bucket = int.from_bytes(digest[:8], "big") / 2**64
    if bucket < 0.8:
        return 0
    if bucket < 0.9:
        return 1
    return 2


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _open_text(path: Path) -> TextIO:
    suffixes = path.suffixes
    if suffixes and suffixes[-1] == ".xz":
        return lzma.open(path, mode="rt", encoding="utf-8", errors="replace")
    if suffixes and suffixes[-1] == ".gz":
        return gzip.open(path, mode="rt", encoding="utf-8", errors="replace")
    return path.open(mode="rt", encoding="utf-8", errors="replace")


def _iter_json_rows(path: Path) -> Iterator[dict[str, object]]:
    with _open_text(path) as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path}:{line_number}: invalid JSON") from exc
            if not isinstance(value, dict):
                raise ValueError(f"{path}:{line_number}: expected a JSON object")
            yield value


def _iter_xlsx_rows(path: Path) -> Iterator[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = [str(value or "") for value in next(rows)]
    except StopIteration as exc:
        workbook.close()
        raise ValueError(f"{path}: empty workbook") from exc
    required = {"lbl", "wght", "dbg", "feat"}
    if not required.issubset(header):
        workbook.close()
        raise ValueError(f"{path}: expected columns {sorted(required)}")
    try:
        for values in rows:
            yield dict(zip(header, values, strict=False))
    finally:
        workbook.close()


def iter_source_rows(paths: Iterable[Path]) -> Iterator[tuple[Path, dict[str, object]]]:
    for path in paths:
        if not path.is_file():
            raise FileNotFoundError(path)
        rows = _iter_xlsx_rows(path) if path.suffix.lower() == ".xlsx" else _iter_json_rows(path)
        for row in rows:
            yield path, row


def parse_record(path: Path, row: Mapping[str, object]) -> ParsedRecord | RejectedRecord:
    raw_label = str(row.get("lbl", "")).lower()
    if raw_label not in {"n", "p"}:
        return RejectedRecord("invalid_label", path.name)
    dbg = str(row.get("dbg", "")).strip()
    script_id = dbg.partition(":")[0].strip()
    if not script_id:
        return RejectedRecord("missing_script_identifier", path.name)

    raw_features = row.get("feat", {})
    if (
        path.suffix.lower() == ".xlsx"
        and isinstance(raw_features, str)
        and len(raw_features) >= EXCEL_CELL_LIMIT
    ):
        return RejectedRecord("excel_cell_limit", path.name)
    features = normalize_feature_mapping(raw_features)
    if features is None:
        return RejectedRecord("invalid_feature_mapping", path.name)
    try:
        weight = float(row.get("wght", 1.0))
    except (TypeError, ValueError):
        return RejectedRecord("invalid_weight", path.name)
    if not np.isfinite(weight) or weight <= 0:
        return RejectedRecord("invalid_weight", path.name)

    return ParsedRecord(
        label=int(raw_label == "p"),
        weight=weight,
        dbg=dbg,
        script_id=script_id,
        features=features,
        feature_hash=canonical_feature_hash(features),
        source_file=str(row.get("source_file") or path.name),
    )


def unique_indexes(indexes: np.ndarray, feature_hashes: np.ndarray) -> np.ndarray:
    seen: set[str] = set()
    selected: list[int] = []
    for index in indexes:
        feature_hash = str(feature_hashes[index])
        if feature_hash in seen:
            continue
        seen.add(feature_hash)
        selected.append(int(index))
    return np.asarray(selected, dtype=np.int64)


def capped_duplicate_indexes(
    indexes: np.ndarray,
    feature_hashes: np.ndarray,
    maximum_per_hash: int,
) -> np.ndarray:
    retained: Counter[str] = Counter()
    selected: list[int] = []
    for index in indexes:
        feature_hash = str(feature_hashes[index])
        if retained[feature_hash] >= maximum_per_hash:
            continue
        retained[feature_hash] += 1
        selected.append(int(index))
    return np.asarray(selected, dtype=np.int64)


def unseen_unique_indexes(
    indexes: np.ndarray,
    feature_hashes: np.ndarray,
    reference_indexes: np.ndarray,
) -> np.ndarray:
    reference = {str(value) for value in feature_hashes[reference_indexes]}
    unseen = np.asarray(
        [int(index) for index in indexes if str(feature_hashes[index]) not in reference],
        dtype=np.int64,
    )
    return unique_indexes(unseen, feature_hashes)


def classification_metrics(
    labels: np.ndarray,
    scores: np.ndarray,
    threshold: float,
) -> dict[str, object]:
    predicted = scores >= threshold
    tn, fp, fn, tp = confusion_matrix(labels, predicted, labels=[0, 1]).ravel()
    return {
        "rows": int(labels.size),
        "positive_rows": int(labels.sum()),
        "threshold": float(threshold),
        "precision": float(precision_score(labels, predicted, zero_division=0)),
        "recall": float(recall_score(labels, predicted, zero_division=0)),
        "f1": float(f1_score(labels, predicted, zero_division=0)),
        "roc_auc": float(roc_auc_score(labels, scores)),
        "pr_auc": float(average_precision_score(labels, scores)),
        "confusion": {"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp)},
    }


def hybrid_classification_metrics(
    labels: np.ndarray,
    scores: np.ndarray,
    threshold: float,
    security_signals: np.ndarray,
) -> dict[str, object]:
    model_positive = scores >= threshold
    combined_positive = model_positive | security_signals
    tn, fp, fn, tp = confusion_matrix(
        labels,
        combined_positive,
        labels=[0, 1],
    ).ravel()
    return {
        "rows": int(labels.size),
        "positive_rows": int(labels.sum()),
        "threshold": float(threshold),
        "precision": float(
            precision_score(labels, combined_positive, zero_division=0)
        ),
        "recall": float(recall_score(labels, combined_positive, zero_division=0)),
        "f1": float(f1_score(labels, combined_positive, zero_division=0)),
        "confusion": {"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp)},
        "model_positive_rows": int(model_positive.sum()),
        "security_signal_rows": int(security_signals.sum()),
        "combined_positive_rows": int(combined_positive.sum()),
    }


def best_f1_threshold(labels: np.ndarray, scores: np.ndarray) -> float:
    best_score = -1.0
    best_threshold = 0.5
    for threshold in np.unique(scores):
        score = f1_score(labels, scores >= threshold, zero_division=0)
        if score > best_score:
            best_score = float(score)
            best_threshold = float(threshold)
    return best_threshold


def target_recall_threshold(
    labels: np.ndarray,
    scores: np.ndarray,
    target_recall: float,
) -> float:
    """Return the highest observed threshold that still meets target recall."""

    candidates = np.unique(scores)
    eligible = [
        float(threshold)
        for threshold in candidates
        if recall_score(labels, scores >= threshold, zero_division=0) >= target_recall
    ]
    return max(eligible) if eligible else float(candidates.min())


def build_dataset(
    inputs: list[Path],
    vocab_size: int,
    min_token_count: int,
    max_rows: int | None,
) -> tuple[sparse.csr_matrix, dict[str, np.ndarray], dict[str, int], dict[str, object]]:
    token_counts: Counter[str] = Counter()
    rejected: Counter[str] = Counter()
    label_counts: Counter[str] = Counter()
    split_counts: Counter[str] = Counter()
    parsed_rows = 0

    for path, row in iter_source_rows(inputs):
        record = parse_record(path, row)
        if isinstance(record, RejectedRecord):
            rejected[record.reason] += 1
            continue
        split = split_for_script(record.script_id)
        if split == 0:
            token_counts.update(record.features)
        label_counts["p" if record.label else "n"] += 1
        split_counts[SPLIT_NAMES[split]] += 1
        parsed_rows += 1
        if max_rows is not None and parsed_rows >= max_rows:
            break

    vocabulary_items = sorted(
        (
            (token, count)
            for token, count in token_counts.items()
            if count >= min_token_count
        ),
        key=lambda item: (-item[1], item[0]),
    )[:vocab_size]
    vocabulary = {
        token: index for index, (token, _count) in enumerate(vocabulary_items)
    }
    if len(vocabulary) != vocab_size:
        raise RuntimeError(
            f"only {len(vocabulary)} terms met min_token_count={min_token_count}"
        )

    data: list[float] = []
    column_indexes: list[int] = []
    row_pointer = [0]
    labels: list[int] = []
    weights: list[float] = []
    groups: list[int] = []
    scripts: list[str] = []
    feature_hashes: list[str] = []
    matched_tokens: list[int] = []
    security_signals: list[bool] = []
    parsed_rows = 0

    for path, row in iter_source_rows(inputs):
        record = parse_record(path, row)
        if isinstance(record, RejectedRecord):
            continue
        matched = 0
        for token, count in record.features.items():
            column = vocabulary.get(token)
            if column is None:
                continue
            column_indexes.append(column)
            data.append(count)
            matched += int(count)
        row_pointer.append(len(data))
        labels.append(record.label)
        weights.append(record.weight)
        groups.append(split_for_script(record.script_id))
        scripts.append(record.script_id)
        feature_hashes.append(record.feature_hash)
        matched_tokens.append(matched)
        security_signals.append(
            any(token.startswith("sec_pair_") for token in record.features)
        )
        parsed_rows += 1
        if max_rows is not None and parsed_rows >= max_rows:
            break

    matrix = sparse.csr_matrix(
        (
            np.asarray(data, dtype=np.float32),
            np.asarray(column_indexes, dtype=np.int32),
            np.asarray(row_pointer, dtype=np.int64),
        ),
        shape=(parsed_rows, vocab_size),
        dtype=np.float32,
    )
    arrays = {
        "labels": np.asarray(labels, dtype=np.int8),
        "weights": np.asarray(weights, dtype=np.float64),
        "groups": np.asarray(groups, dtype=np.int8),
        "scripts": np.asarray(scripts),
        "feature_hashes": np.asarray(feature_hashes),
        "matched_tokens": np.asarray(matched_tokens, dtype=np.int32),
        "security_signals": np.asarray(security_signals, dtype=bool),
    }
    audit = {
        "parsed_rows": parsed_rows,
        "label_counts": dict(label_counts),
        "split_counts": dict(split_counts),
        "rejected_rows": dict(rejected),
        "zero_vocabulary_coverage_rows": int((arrays["matched_tokens"] == 0).sum()),
        "unique_scripts": int(np.unique(arrays["scripts"]).size),
        "input_files": [
            {"path": path.name, "size": path.stat().st_size, "sha256": sha256_file(path)}
            for path in inputs
        ],
    }
    return matrix, arrays, vocabulary, audit


def fit_and_select(
    matrix: sparse.csr_matrix,
    arrays: dict[str, np.ndarray],
    target_recall: float,
    max_train_duplicates: int,
) -> tuple[LGBMClassifier, dict[str, object], dict[str, object]]:
    labels = arrays["labels"]
    groups = arrays["groups"]
    hashes = arrays["feature_hashes"]
    eligible = arrays["matched_tokens"] > 0

    hash_labels: dict[str, set[int]] = defaultdict(set)
    for feature_hash, label in zip(hashes[eligible], labels[eligible], strict=True):
        hash_labels[str(feature_hash)].add(int(label))
    conflicting = {
        feature_hash for feature_hash, values in hash_labels.items() if len(values) > 1
    }
    if conflicting:
        eligible &= np.asarray(
            [str(feature_hash) not in conflicting for feature_hash in hashes],
            dtype=bool,
        )

    train_all = np.flatnonzero(eligible & (groups == 0))
    validation_all = np.flatnonzero(eligible & (groups == 1))
    test_all = np.flatnonzero(eligible & (groups == 2))
    train = capped_duplicate_indexes(
        train_all,
        hashes,
        max_train_duplicates,
    )
    validation = unseen_unique_indexes(validation_all, hashes, train)
    train_validation_all = np.concatenate([train_all, validation_all])
    train_validation = capped_duplicate_indexes(
        train_validation_all,
        hashes,
        max_train_duplicates,
    )
    test = unseen_unique_indexes(test_all, hashes, train_validation)

    if min(labels[validation].sum(), labels[test].sum()) < 20:
        raise RuntimeError("strict validation/test splits contain too few positive rows")

    candidate_results: list[dict[str, object]] = []
    best_model: LGBMClassifier | None = None
    best_result: dict[str, object] | None = None
    for candidate in DEFAULT_CANDIDATES:
        model = LGBMClassifier(
            objective="binary",
            n_estimators=2_000,
            subsample=0.8,
            colsample_bytree=0.8,
            reg_alpha=0.1,
            reg_lambda=1.0,
            class_weight="balanced",
            random_state=RANDOM_STATE,
            n_jobs=-1,
            verbosity=-1,
            **candidate,
        )
        model.fit(
            matrix[train],
            labels[train],
            sample_weight=arrays["weights"][train],
            eval_set=[(matrix[validation], labels[validation])],
            eval_metric="average_precision",
            callbacks=[early_stopping(100, verbose=False), log_evaluation(0)],
        )
        validation_scores = model.predict_proba(matrix[validation])[:, 1]
        f1_threshold = best_f1_threshold(labels[validation], validation_scores)
        recall_threshold = target_recall_threshold(
            labels[validation],
            validation_scores,
            target_recall,
        )
        result = {
            "parameters": candidate,
            "best_iteration": int(model.best_iteration_),
            "validation": classification_metrics(
                labels[validation], validation_scores, recall_threshold
            ),
            "validation_best_f1": classification_metrics(
                labels[validation], validation_scores, f1_threshold
            ),
            "validation_at_0_5": classification_metrics(
                labels[validation], validation_scores, 0.5
            ),
        }
        candidate_results.append(result)
        if (
            best_result is None
            or result["validation"]["pr_auc"] > best_result["validation"]["pr_auc"]  # type: ignore[index]
        ):
            best_model = model
            best_result = result

    if best_model is None or best_result is None:
        raise RuntimeError("model selection did not produce a candidate")

    threshold = float(best_result["validation"]["threshold"])  # type: ignore[index]
    test_scores = best_model.predict_proba(matrix[test])[:, 1]
    evaluation = {
        "protocol": (
            "Deterministic script-level split; train-only vocabulary; invalid, "
            "zero-coverage and conflicting-label feature bags excluded; training "
            f"duplicates capped at {max_train_duplicates} rows per feature bag; "
            "validation and test are unique and unseen by earlier splits; "
            f"deployment threshold targets {target_recall:.1%} validation recall."
        ),
        "feature_contract": FEATURE_CONTRACT,
        "conflicting_feature_hashes": len(conflicting),
        "split_rows": {
            "train_capped": int(train.size),
            "validation_unseen_unique": int(validation.size),
            "test_unseen_unique": int(test.size),
        },
        "training_duplicate_cap": max_train_duplicates,
        "candidates": candidate_results,
        "selected": best_result,
        "test": classification_metrics(labels[test], test_scores, threshold),
        "test_at_0_5": classification_metrics(labels[test], test_scores, 0.5),
        "test_hybrid": hybrid_classification_metrics(
            labels[test],
            test_scores,
            threshold,
            arrays["security_signals"][test],
        ),
        "test_hybrid_at_0_5": hybrid_classification_metrics(
            labels[test],
            test_scores,
            0.5,
            arrays["security_signals"][test],
        ),
    }

    final_parameters = {
        **best_result["parameters"],  # type: ignore[arg-type]
        "objective": "binary",
        "n_estimators": int(best_result["best_iteration"]),
        "subsample": 0.8,
        "colsample_bytree": 0.8,
        "reg_alpha": 0.1,
        "reg_lambda": 1.0,
        "class_weight": "balanced",
        "random_state": RANDOM_STATE,
        "n_jobs": -1,
        "verbosity": -1,
    }
    deployment_model = LGBMClassifier(**final_parameters)
    deployment_model.fit(
        matrix[train_validation],
        labels[train_validation],
        sample_weight=arrays["weights"][train_validation],
    )
    deployment_test_scores = deployment_model.predict_proba(matrix[test])[:, 1]
    evaluation["deployment_test_at_0_5"] = classification_metrics(
        labels[test],
        deployment_test_scores,
        0.5,
    )
    evaluation["deployment_test_hybrid_at_0_5"] = hybrid_classification_metrics(
        labels[test],
        deployment_test_scores,
        0.5,
        arrays["security_signals"][test],
    )
    training = {
        "parameters": final_parameters,
        "validation_selected_threshold": threshold,
        "deployment_training_rows_capped": int(train_validation.size),
        "deployment_training_positive_rows": int(labels[train_validation].sum()),
        "maximum_rows_per_training_feature_bag": max_train_duplicates,
    }
    return deployment_model, evaluation, training


def write_artifacts(
    output_root: Path,
    model: LGBMClassifier,
    vocabulary: dict[str, int],
    audit: dict[str, object],
    evaluation: dict[str, object],
    training: dict[str, object],
    args: argparse.Namespace,
) -> None:
    model_dir = output_root / "models"
    preprocessing_dir = output_root / "preprocessing"
    results_dir = output_root / "docs" / "results"
    for directory in (model_dir, preprocessing_dir, results_dir):
        directory.mkdir(parents=True, exist_ok=True)

    pickle_path = model_dir / "lightgbm_security_v2.pkl"
    native_path = model_dir / "lightgbm_security_v2.txt"
    vocabulary_path = preprocessing_dir / "vocab_security_v2.json"
    metadata_path = model_dir / "lightgbm_security_v2_metadata.json"
    evaluation_path = results_dir / "lightgbm_security_v2_evaluation.json"

    joblib.dump(model, pickle_path)
    model.booster_.save_model(native_path)
    vocabulary_path.write_text(
        json.dumps(vocabulary, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    evaluation_path.write_text(
        json.dumps(evaluation, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    metadata = {
        "artifact_version": 3,
        "model_name": "DOM-XSS LightGBM Security v2",
        "model_family": "LightGBM",
        "output_semantics": "risk_score_not_calibrated_probability",
        "feature_contract": FEATURE_CONTRACT,
        "feature_representation": (
            f"normalized AST token counts plus deterministic source/sink "
            f"co-occurrence features, top-{len(vocabulary)} train-only vocabulary"
        ),
        "split_contract": "deterministic 80/10/10 assignment by dbg script identifier",
        "dataset": {
            "name": "DOM XSS Web Vulnerability Dataset",
            "url": args.dataset_url,
            "paper": args.paper_url,
            "audit": audit,
        },
        "training": training,
        "evaluation": {
            "protocol": evaluation["protocol"],
            "test": evaluation["test"],
            "test_at_0_5": evaluation["test_at_0_5"],
            "test_hybrid": evaluation["test_hybrid"],
            "test_hybrid_at_0_5": evaluation["test_hybrid_at_0_5"],
            "deployment_test_at_0_5": evaluation["deployment_test_at_0_5"],
            "deployment_test_hybrid_at_0_5": evaluation[
                "deployment_test_hybrid_at_0_5"
            ],
        },
        "limitations": [
            (
                "This LightGBM derivative is not the paper's TensorFlow DNN and does "
                "not reproduce the paper's full 32-million-function experiment."
            ),
            (
                "When legacy XLSX inputs are used, rows at Excel's 32,767-character "
                "cell limit are rejected because their feature dictionaries are truncated."
            ),
            (
                "Scores rank candidates and are not calibrated probabilities of exploitability."
            ),
        ],
    }
    metadata_path.write_text(
        json.dumps(metadata, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    for path in (pickle_path, native_path, vocabulary_path, metadata_path, evaluation_path):
        print(f"wrote {path}")


def main() -> None:
    args = parse_args()
    if args.vocab_size <= 0 or args.min_token_count <= 0:
        raise ValueError("vocab size and minimum token count must be positive")
    if not 0.0 < args.target_recall <= 1.0:
        raise ValueError("target recall must be in the interval (0, 1]")
    if args.max_train_duplicates <= 0:
        raise ValueError("maximum training duplicates must be positive")
    matrix, arrays, vocabulary, audit = build_dataset(
        args.inputs,
        args.vocab_size,
        args.min_token_count,
        args.max_rows,
    )
    model, evaluation, training = fit_and_select(
        matrix,
        arrays,
        args.target_recall,
        args.max_train_duplicates,
    )
    write_artifacts(
        args.output_root,
        model,
        vocabulary,
        audit,
        evaluation,
        training,
        args,
    )


if __name__ == "__main__":
    main()
